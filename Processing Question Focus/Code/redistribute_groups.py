#!/usr/bin/env python3
"""Redistribute items across 4 Groups + renumber S01-S20 within each group."""
import re, os
from collections import defaultdict

BASE = os.path.dirname(os.path.abspath(__file__))

# ── Build: list of (old_group, old_S, is_cond_14) for all 80 items ──
items = []
# G1: all Cond 1-4 (S01-S20)
for s in range(1, 21):
    items.append((1, s, True))
# G2: S01-S12 Cond 1-4, S13-S20 Cond 5-8
for s in range(1, 13):
    items.append((2, s, True))
for s in range(13, 21):
    items.append((2, s, False))
# G3: all Cond 5-8 (S01-S20)
for s in range(1, 21):
    items.append((3, s, False))
# G4: all Cond 5-8 (S01-S20)
for s in range(1, 21):
    items.append((4, s, False))

print(f'Total items: {len(items)}')
print(f'  Cond 1-4: {sum(1 for _,_,c14 in items if c14)}')
print(f'  Cond 5-8: {sum(1 for _,_,c14 in items if not c14)}')

# ── Assign to new groups ──
# Separate into two pools then interleave
pool_14 = [(og, s) for og, s, c14 in items if c14]
pool_58 = [(og, s) for og, s, c14 in items if not c14]

# Assign to groups in round-robin, preserving original order within pool
new_groups = {1: [], 2: [], 3: [], 4: []}

for i, (og, s) in enumerate(pool_14):
    ng = (i % 4) + 1
    new_groups[ng].append((og, s, True))

for i, (og, s) in enumerate(pool_58):
    ng = (i % 4) + 1
    new_groups[ng].append((og, s, False))

# Renumber S within each new group: S01-S20
# Keep Cond 1-4 items first, then Cond 5-8
old_to_new = {}  # (old_group, old_S) -> (new_group, new_S)

for ng in [1, 2, 3, 4]:
    items_in_group = new_groups[ng]
    # Sort: Cond 1-4 first, then Cond 5-8; within each, by original group then S
    items_in_group.sort(key=lambda x: (0 if x[2] else 1, x[0], x[1]))
    for new_s, (og, old_s, c14) in enumerate(items_in_group, 1):
        old_to_new[(og, old_s)] = (ng, f'S{new_s:02d}')
        cond_type = '1-4' if c14 else '5-8'
        print(f'  Old G{og}_S{old_s:02d} → New G{ng}_S{new_s:02d}  ({cond_type})')
    print(f'  --- Group {ng}: {len(items_in_group)} items ---')

# Verify
from collections import Counter
gc = Counter(ng for ng, _ in old_to_new.values())
print(f'\nVerification: {dict(sorted(gc.items()))}')
assert gc[1] == 20 and gc[2] == 20 and gc[3] == 20 and gc[4] == 20

# Check unique labels
new_labels = set()
for (og, old_s), (ng, new_s) in old_to_new.items():
    new_labels.add(f'Group_{ng}_S{new_s}')
assert len(new_labels) == 80, f'Expected 80 unique labels, got {len(new_labels)}'
print('All 80 new labels are unique ✅')

# ── Apply to all 4 xlsx files ──
import openpyxl

# Build regex replacement mapping
# Map: "Group_OLDG_Condition_COND_SOLDS" -> "Group_NEWG_Condition_COND_NEWS"
def build_replacement_map(list_num):
    """For a given list, build the replacement mapping for all items."""
    # We need to read each item's current label, look up old (group, S),
    # then generate new label with new (group, S) but keeping Condition
    mapping = {}
    for (og, old_s), (ng, new_s) in old_to_new.items():
        # In List list_num, the condition for this item depends on the rotation
        # Cond 1-4 items rotate: (list 1→cond1, 2→cond2, 3→cond3, 4→cond4)
        # Cond 5-8 items rotate: (list 1→cond5, 2→cond6, 3→cond7, 4→cond8)
        # But the exact condition depends on the item's position in the rotation

        # Actually, we don't need to compute the condition.
        # We just need to find the old label in the xlsx and replace it.
        # The old label has the form: Group_OLDG_Condition_COND_SOLDS
        # We need to match this pattern and replace Group + S.

        # We'll do this by regex matching in the actual file content
        pass
    return mapping

# Simpler approach: use regex to find each label and transform
def transform_label(match, old_to_new):
    """Transform a matched label using old_to_new mapping."""
    full = match.group(0)
    parts = full.split('_')
    # parts: ['Group', 'X', 'Condition', 'Y', 'SZZ']
    old_g = int(parts[1])
    cond = parts[3]  # keep
    old_s_str = parts[4]  # S01, S02, etc.
    old_s = int(old_s_str[1:])

    key = (old_g, old_s)
    if key in old_to_new:
        new_g, new_s_str = old_to_new[key]
        parts[1] = str(new_g)
        parts[4] = new_s_str
        return '_'.join(parts)
    return full

total_changes = 0
for ln in range(1, 5):
    path = os.path.join(BASE, f'IBEX_Maze_List{ln}.xlsx')
    wb = openpyxl.load_workbook(path)
    ws = wb['Template']
    changes = 0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cell = row[0]
        old_text = str(cell.value) if cell.value else ''
        new_text = re.sub(
            r'Group_\d+_Condition_\d+_S\d+',
            lambda m: transform_label(m, old_to_new),
            old_text
        )
        if old_text != new_text:
            changes += 1
            cell.value = new_text

    wb.save(path)
    print(f'List {ln}: {changes} labels changed')
    total_changes += changes

print(f'\nTotal: {total_changes} label changes across all 4 xlsx files ✅')
