#!/usr/bin/env python3
"""
Convert Excel Maze experiment data to PCIbex-compatible CSV table.

Reads:
  - Maze_with Context_Template.xlsx  (practice items, 16 rows)
  - IBEX_Maze_List1.xlsx through List4.xlsx (experimental items, 80 rows each)

Outputs:
  - items.csv  — single table with all items for PCIbex
"""

import openpyxl
import re
import csv
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ── helpers ──────────────────────────────────────────────────

def parse_js_like(text):
    """
    Parse a JavaScript-like array string from Excel.
    Returns dict with keys: label, html, maze_s, maze_a, question
    """
    if not text or text.strip() == '':
        return None

    # Extract the label (first quoted string in brackets)
    label_m = re.search(r'\["([^"]+)"', text)
    if not label_m:
        return None
    label = label_m.group(1)

    # Extract html from Message (handles "..." + "..." concatenation)
    html = ""
    html_start = text.find('html:')
    if html_start >= 0:
        # Find the closing brace of the Message object
        # The html value may contain "..." + "..." concatenation
        # We extract everything between html: and the next unquoted }, or Maze marker
        rest = text[html_start+5:]  # skip 'html:'
        # Collect all "..." strings connected by +
        html_parts = []
        pos = 0
        while pos < len(rest):
            # Skip whitespace
            while pos < len(rest) and rest[pos] in ' \t':
                pos += 1
            if pos < len(rest) and rest[pos] == '"':
                # Extract quoted string
                end_pos = pos + 1
                while end_pos < len(rest):
                    if rest[end_pos] == '\\' and end_pos+1 < len(rest):
                        end_pos += 2  # skip escaped char
                    elif rest[end_pos] == '"':
                        break
                    else:
                        end_pos += 1
                if end_pos < len(rest):
                    html_parts.append(rest[pos+1:end_pos])
                    pos = end_pos + 1
                else:
                    break
            elif pos < len(rest) and rest[pos] == '+':
                pos += 1  # continue to next string
            elif pos < len(rest) and rest[pos] == '}':
                break  # end of Message object
            else:
                break
        html = ''.join(html_parts)

    # Extract Maze s and a
    s_m = re.search(r's:"((?:[^"\\]|\\.)*)"', text)
    a_m = re.search(r'a:"((?:[^"\\]|\\.)*)"', text)
    maze_s = s_m.group(1) if s_m else ""
    maze_a = a_m.group(1) if a_m else ""

    # Extract Question
    q_m = re.search(r'q:\s*"((?:[^"\\]|\\.)*)"', text)
    question = q_m.group(1) if q_m else ""

    return {
        'label': label,
        'html': html,
        'maze_s': maze_s,
        'maze_a': maze_a,
        'question': question,
    }


def parse_label(label):
    """Extract group, condition, and item number from label like 'Group_2_Condition_5_S13'"""
    m = re.match(r'Group_(\d+)_Condition_(\d+)_S(\d+)', label)
    if m:
        return int(m.group(1)), int(m.group(2)), int(m.group(3))
    return None, None, None


def determine_answer(question, context_html):
    """
    Determine expected answer for comprehension question.
    In this experiment, all questions are True/False statements.
    We default to marking 'True' as the correct answer -
    the actual answer depends on whether the question statement
    matches the context. Since we can't automatically determine this,
    we mark questions where the statement contradicts the context as 'False'.

    Actually: we can't reliably auto-determine this. We'll set all to 'True'
    and the user can adjust. OR we can try heuristic matching.

    For now: we leave expected_answer blank and let the experimenter fill.
    Or better: on pcibex.net, the experimenter can set correct answers there.
    """
    # Simple heuristic: if question contains negation markers like 不/没/别/否,
    # it might be False when the context is positive
    # But this is unreliable — we'll leave it for manual checking
    return ""


def process_excel(filepath, item_type, list_num=None):
    """Process one Excel file and return list of item dicts."""
    items = []
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Template']

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        text = str(row[0]) if row[0] else ''
        parsed = parse_js_like(text)
        if not parsed:
            continue

        group, condition, item_num = parse_label(parsed['label'])

        items.append({
            'label': parsed['label'],
            'group': group if group else '',
            'condition': condition if condition else '',
            'item_num': item_num if item_num else '',
            'type': item_type,
            'list': list_num if list_num else '',
            'message_html': parsed['html'],
            'maze_s': parsed['maze_s'],
            'maze_a': parsed['maze_a'],
            'question': parsed['question'],
        })

    return items


# ── main ────────────────────────────────────────────────────

def main():
    all_items = []

    # 1. Practice items (template)
    practice_path = os.path.join(BASE_DIR, 'Maze_with Context_Template.xlsx')
    if os.path.exists(practice_path):
        practice_items = process_excel(practice_path, 'practice')
        all_items.extend(practice_items)
        print(f"  Practice: {len(practice_items)} items")

    # 2. Experimental items (4 lists)
    for list_num in range(1, 5):
        list_path = os.path.join(BASE_DIR, f'IBEX_Maze_List{list_num}.xlsx')
        if os.path.exists(list_path):
            list_items = process_excel(list_path, 'experiment', list_num)
            all_items.extend(list_items)
            print(f"  List {list_num}: {len(list_items)} items")

    fieldnames = [
        'label', 'group', 'condition', 'item_num', 'type', 'list',
        'message_html', 'maze_s', 'maze_a', 'question'
    ]

    # 3. Write combined CSV (all items)
    csv_path = os.path.join(BASE_DIR, 'items.csv')
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for item in all_items:
            writer.writerow(item)
    print(f"\nTotal: {len(all_items)} items written to {csv_path}")

    # 4. Write separate CSV files for PCIbex
    practice_items = [i for i in all_items if i['type'] == 'practice']
    experiment_items = [i for i in all_items if i['type'] == 'experiment']

    # Practice only
    practice_path = os.path.join(BASE_DIR, 'practice.csv')
    with open(practice_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for item in practice_items:
            writer.writerow(item)
    print(f"  practice.csv: {len(practice_items)} items")

    # Each list (experimental items only)
    for ln in range(1, 5):
        list_items = [i for i in experiment_items if i['list'] == ln]
        list_path = os.path.join(BASE_DIR, f'list{ln}.csv')
        with open(list_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for item in list_items:
                writer.writerow(item)
        print(f"  list{ln}.csv: {len(list_items)} experimental items")

    # 5. Print summary
    practice = [i for i in all_items if i['type'] == 'practice']
    experiment = [i for i in all_items if i['type'] == 'experiment']
    print(f"\nSummary:")
    print(f"  Practice trials:     {len(practice)}")
    print(f"  Experimental trials: {len(experiment)}")
    for ln in range(1, 5):
        n = len([i for i in experiment if i['list'] == ln])
        print(f"    List {ln}: {n} experimental trials")


if __name__ == '__main__':
    main()
