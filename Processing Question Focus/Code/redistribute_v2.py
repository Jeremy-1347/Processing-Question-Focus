#!/usr/bin/env python3
"""Redistribute items across 4 Groups with balanced rotation patterns."""
import re, os, csv
from collections import defaultdict
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))

# ── Read current data to get rotation patterns ──
all_rows = []
for ln in range(1, 5):
    with open(f'list{ln}.csv', encoding='utf-8-sig') as f:
        for row in csv.DictReader(f):
            all_rows.append(row)

item_pattern = {}
item_type = {}
for r in all_rows:
    if r['item_id'] not in item_pattern:
        item_pattern[r['item_id']] = {}
    item_pattern[r['item_id']][r['list']] = int(r['condition'])

for item_id, conds in item_pattern.items():
    item_pattern[item_id] = tuple(conds.get(str(ln), 0) for ln in range(1, 5))
    item_type[item_id] = '14' if all(c <= 4 for c in item_pattern[item_id]) else '58'

# Group by rotation pattern
patterns_14 = defaultdict(list)
patterns_58 = defaultdict(list)
for item_id, pattern in item_pattern.items():
    if item_type[item_id] == '14':
        patterns_14[pattern].append(item_id)
    else:
        patterns_58[pattern].append(item_id)

# ── Redistribute: each new group gets equal from each rotation pattern ──
new_assignment = defaultdict(list)  # ng -> [item_ids]
for pattern, items in sorted(patterns_14.items()):
    for i, item_id in enumerate(items):
        ng = (i % 4) + 1
        new_assignment[ng].append(item_id)
for pattern, items in sorted(patterns_58.items()):
    for i, item_id in enumerate(items):
        ng = (i % 4) + 1
        new_assignment[ng].append(item_id)

# Sort items within each group: Cond1-4 first, then Cond5-8
for ng in [1, 2, 3, 4]:
    new_assignment[ng].sort(key=lambda iid: (0 if item_type[iid] == '14' else 1, iid))

# ── Build mapping: item_id -> (new_group, new_S) ──
item_to_new = {}
for ng in [1, 2, 3, 4]:
    for new_s, item_id in enumerate(new_assignment[ng], 1):
        item_to_new[item_id] = (str(ng), f'S{new_s:02d}')

# ── Also build (old_group, old_S) -> (new_group, new_S) for label replacement ──
# The old label format in the xlsx: Group_OLDG_Condition_COND_SOLDS
# We need to map (old_group, old_S) from item_id (e.g., "G1_S05" -> old_g=1, old_s=5)
old_to_new = {}
for r in all_rows:
    label = r['label']
    item_id = r['item_id']
    if item_id in item_to_new:
        # Extract old group and S from the CURRENT label (which is already modified once)
        parts = label.split('_')
        old_g = int(parts[1])
        old_s_str = parts[4]
        old_s = int(old_s_str[1:])
        key = (old_g, old_s, item_id)  # use item_id to track across lists
        old_to_new[(old_g, old_s)] = item_to_new[item_id]

# ── Apply to xlsx files ──
# Restore original xlsx first
import shutil
for ln in range(1, 5):
    src = f'/Users/jeremyp./Desktop/实验2/普通话实验二/IBEX_Maze_List{ln}.xlsx'
    dst = os.path.join(BASE, f'IBEX_Maze_List{ln}.xlsx')
    shutil.copy(src, dst)
print('Restored original xlsx files')

# Apply label changes
def transform_label(match, item_to_new, current_list):
    """Transform label: find the item_id via the label's condition, then apply new Group+S."""
    # We need a way to identify which item this is.
    # The label is: Group_X_Condition_Y_SZZ
    # We can't directly know the item_id from just the label because we changed it before.
    # But we restored from originals, so the label is the ORIGINAL label.
    # Original: Group_OG_Condition_C_SOS
    # item_id = G{OG}_S{OS}
    full = match.group(0)
    parts = full.split('_')
    old_g = parts[1]
    old_S = parts[4]  # e.g., "S05"
    item_id = f'G{old_g}_{old_S}'

    if item_id in item_to_new:
        new_g, new_S = item_to_new[item_id]
        parts[1] = new_g
        parts[4] = new_S
        return '_'.join(parts)
    return full

total = 0
for ln in range(1, 5):
    path = os.path.join(BASE, f'IBEX_Maze_List{ln}.xlsx')
    wb = openpyxl.load_workbook(path)
    ws = wb['Template']
    changes = 0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cell = row[0]
        old_text = str(cell.value) if cell.value else ''
        new_text = re.sub(
            r'Group_\d+_Condition_\d+_S\d+',
            lambda m: transform_label(m, item_to_new, ln),
            old_text
        )
        if old_text != new_text:
            changes += 1
            cell.value = new_text

    wb.save(path)
    print(f'List {ln}: {changes} labels changed')
    total += changes

print(f'\nTotal: {total} label changes')
PYEOF
